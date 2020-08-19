import cv2
import numpy as np
import operator
import timeit
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, named_styles, colors, Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

MAX_CORNERS = 100000

# test -2, test5, test6

scale1 = 30
scale2 = 30
operation1 = 2
operation2 = 1
operation3 = 1
operation4 = 12

checkline1 = 10
checkline2 = 10

# getmrpointlist

line_thick = 2

# ver hough
vertical_quality = 0.01
vertical_threshold = 1
ver_minline = 2
ver_maxgap = 6
# hor hough
horizontal_quality = 0.01
horizontal_threshold = 10
hor_minline = 2
hor_maxgap = 1

corner_distance = 5
align_range = 10
distance_square = 100


# # test3
#
# scale1 = 30
# scale2 = 30
# operation1 = 2
# operation2 = 1
# operation3 = 1
# operation4 = 12
#
# checkline1 = 10
# checkline2 = 10
#
# # getmrpointlist
#
# line_thick = 3
#
# # ver hough
# vertical_quality = 0.1
# vertical_threshold = 100
# ver_minline = 30
# ver_maxgap = 10
# # hor hough
# horizontal_quality = 0.1
# horizontal_threshold = 100
# hor_minline = 30
# hor_maxgap = 10
#
# corner_distance = 15
# align_range = 10
# distance_square = 100


# #
# # test4
#
# scale1 = 30
# scale2 = 30
# operation1 = 2
# operation2 = 1
# operation3 = 1
# operation4 = 12
#
# checkline1 = 10
# checkline2 = 10
#
# # getmrpointlist
#
# line_thick = 2
#
# # ver hough
# vertical_quality = 1
# vertical_threshold = 30
# ver_minline = 20
# ver_maxgap = 10
# # hor hough
# horizontal_quality = 1
# horizontal_threshold = 30
# hor_minline = 40
# hor_maxgap = 10
#
# corner_distance = 25
# align_range = 10
# distance_square = 100



def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws[cell_range.split(":")[0]]

    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment
        rows = ws[cell_range]

    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top

    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right

        if fill:
            for c in row:
                c.fill = fill


def openOperation(img, imgsize, select):  # 열림 연산 // select : 0(vertical), 1(horizontal)
    # 이미지 크기에 따른 scale 값 지정
    if imgsize > 800:
        scale = 30
    else:
        scale = 30

    size = imgsize / scale
    if select == 0:
        window = np.ones((int(size), 1), np.uint8)
    elif select == 1:
        window = np.ones((1, int(size)), np.uint8)

    # para1 : 이미지, para2 : 커널, para3 : erode 반복 횟수
    img = cv2.erode(img, window, iterations=1)
    # para1 : 이미지, para2 : 커널, para3 : dilate 반복 횟수
    img = cv2.dilate(img, window, iterations=1)
    return img

def openOperation2(img, select):  # 열림 연산 // select : 0(vertical), 1(horizontal)
    if select == 0:
        img = cv2.dilate(img, np.ones((operation1, operation2), np.uint8), iterations=1)
    elif select == 1:
        img = cv2.dilate(img, np.ones((operation3, operation4), np.uint8), iterations=1)

    return img


def CheckLine(horizontal_img, x1, x2, y):
    xpos_avg = int((x1 + x2) / 2)  # 사각형의 x1, x2의 평균
    maxim = checkline1
    half = int(maxim/2)
    for i in range(maxim):
        if horizontal_img[int(y) - half + i][xpos_avg] > 240:  # 240보다 크면 선이 있다고 판단한다.
            return True

    return False


def CheckLine2(vertical_img, y1, y2, x):
    ypos_avg = int((y1 + y2) / 2)  # 사각형의 x1, x2의 평균
    maxim = checkline2
    half = int(maxim/2)
    for i in range(maxim):
        if vertical_img[ypos_avg][int(x) - half + i] > 240:  # 240보다 크면 선이 있다고 판단한다.
            return True

    return False


def GetBlockList(gray, rows, cols):
    R_vertical = np.zeros((rows, cols,3), dtype=np.uint8)
    R_horizontal = np.zeros((rows,cols,3), dtype=np.uint8)
    R_vertical[:]=(255,255,255)
    R_horizontal[:]=(255,255,255)
    result = np.zeros((rows, cols, 3), dtype=np.uint8)
    result[:] = (255, 255, 255)

    vertical = openOperation(gray, rows, 0)
    horizontal = openOperation(gray, cols, 1)

    vertical = openOperation2(vertical, 0)
    horizontal = openOperation2(horizontal, 1)

    # vertical = cv2.erode(vertical, np.ones((1, 2), np.uint8), iterations=1)
    horizontal = cv2.dilate(horizontal, np.ones((2, 1), np.uint8), iterations=1)

    ver_threshold = vertical_threshold
    vertical_lines = cv2.HoughLinesP(vertical, vertical_quality, np.pi / 180, ver_threshold, None, ver_minline, ver_maxgap)
    hor_threshold = horizontal_threshold
    horizontal_lines = cv2.HoughLinesP(horizontal, horizontal_quality, np.pi / 180, hor_threshold, None, hor_minline, hor_maxgap)

    for i in range(len(vertical_lines)):
        cv2.line(R_vertical, (vertical_lines[i][0][0], vertical_lines[i][0][1]), (vertical_lines[i][0][2], vertical_lines[i][0][3]), (0, 0, 0), line_thick)
    for i in range(len(horizontal_lines)):
        cv2.line(R_horizontal, (horizontal_lines[i][0][0], horizontal_lines[i][0][1]),
                 (horizontal_lines[i][0][2], horizontal_lines[i][0][3]), (0, 0, 0), line_thick)

    sec = R_vertical + R_horizontal

    g_sec = cv2.cvtColor(sec,cv2.COLOR_RGB2GRAY)
    CORNERS = cv2.goodFeaturesToTrack(g_sec, MAX_CORNERS, 0.01, corner_distance, 10)  # 코너 추출 함수

    xscore = [0] * len(CORNERS)
    yscore = [0] * len(CORNERS)

    R = [0] * len(CORNERS)
    for i in range(len(CORNERS)):
        for j in range(i, len(CORNERS)):
            if abs(CORNERS[i][0][0] - CORNERS[j][0][0]) < align_range and i != j:
                CORNERS[j][0][0] = CORNERS[i][0][0]

            if abs(CORNERS[i][0][1] - CORNERS[j][0][1]) < align_range and i != j:
                CORNERS[j][0][1] = CORNERS[i][0][1]

    for i in range(len(CORNERS)):
        for j in range(len(CORNERS)):
            if CORNERS[i][0][0]==CORNERS[j][0][0] and i !=j:
                if CORNERS[i][0][1] != CORNERS[j][0][1]:
                    xscore[i] = xscore[i] + 1

    for i in range(len(CORNERS)):
        for j in range(len(CORNERS)):
            if CORNERS[i][0][1] == CORNERS[j][0][1] and i != j:
                if CORNERS[i][0][0] != CORNERS[j][0][0]:
                    yscore[i] = yscore[i] + 1

    for i in range(len(CORNERS)):
        for j in range(i, len(CORNERS)):
            if (pow(CORNERS[i][0][0]-CORNERS[j][0][0], 2) + pow(CORNERS[i][0][1]-CORNERS[j][0][1], 2)) < distance_square:
                if i != j:
                    if xscore[i] > xscore[j]:
                        R[j] = -1
                    else:
                        R[i] = -1
                    if yscore[i] > yscore[j]:
                        R[j] = -1
                    else:
                        R[i] = -1
    #
    #for i in range(len(CORNERS)):
    #    for j in range(len(CORNERS)):
    #        if CORNERS[i][0][0] == CORNERS[j][0][0] or CORNERS[i][0][1] == CORNERS[j][0][1]:
    #            if xscore[i] > 0 and yscore[i] > 0 and xscore[j] > 0 and yscore[j] > 0 and R[i] == 0 and R[j] == 0:
    #                if abs(CORNERS[i][0][0]-CORNERS[j][0][0]) < 2500 and abs(CORNERS[i][0][1] - CORNERS[j][0][1]) < 500:
    #                    cv2.line(result, (CORNERS[i][0][0], CORNERS[i][0][1]),
    #                             (CORNERS[j][0][0], CORNERS[j][0][1]), (0, 0, 255), 2)
    #end = timeit.default_timer()

    #print("Execution time > {:.2f}s\n".format(end - start))

    real_corner = []
    for i in range(len(CORNERS)):
        if xscore[i] > 0 and yscore[i] > 0 and R[i] == 0:
            real_corner.append([CORNERS[i][0][0], CORNERS[i][0][1], 0, 0])
            # x, y = CORNERS[i].ravel()
            # cv2.circle(result, (x, y), 5, 0, 1)  # 그냥 보는 용도

    # result = cv2.resize(result, None, fx=0.7, fy=0.7, interpolation=cv2.INTER_AREA)
    vertical = cv2.dilate(vertical, np.ones((1, 2), np.uint8), iterations=1)

    #cv2.imshow('result', result)
    # cv2.imshow('v', vertical)
    # cv2.imshow('h', horizontal)
    real_corner.sort(key=lambda corner: corner[0])

    arr = []
    xnum = 0
    ynum = 0
    pre = real_corner[0][0]

    for i in range(len(real_corner)):
        if pre != real_corner[i][0]:
            xnum = xnum + 1
        real_corner[i][2] = xnum
        pre = real_corner[i][0]

    real_corner.sort(key=lambda corner: (corner[1], corner[0]))
    pre = real_corner[0][1]

    for i in range(len(real_corner)):
        if pre != real_corner[i][1]:
            ynum = ynum + 1
        real_corner[i][3] = ynum
        arr.append(real_corner[i])
        pre = real_corner[i][1]

    arr.sort(key=operator.itemgetter(2))
    arr.sort(key=operator.itemgetter(3))

    #print(arr)

    xnum = xnum + 1
    ynum = ynum + 1
    #print(xnum, ":", ynum)

    block_point = []  # 블록 좌표가 저장되는 곳. [x1, y1, x2, y2] (매핑 좌표)

    next = 1
    visit = [False]*len(arr)
    i = 0
    while i < len(arr):
        if visit[i] or arr[i][2] == xnum - 1 or arr[i+1][2] < arr[i][2] or (CheckLine(horizontal, arr[i][0], arr[i+1][0], arr[i][1]) == False):
            visit[i] = True
            i += 1
            continue
        prev_x = arr[i][0]
        j = i
        while j < len(arr):
            if arr[i+next][2] == arr[j][2] and i+next < j:
                if CheckLine2(vertical, arr[i+next][1], arr[j][1], arr[j][0]) == False:
                    j = i
                    visit[i+next] = True
                    next = next + 1
                    continue
                elif(CheckLine(horizontal, prev_x, arr[j][0], arr[j][1])):
                    block_point.append([arr[i][2], arr[i][3], arr[j][2], arr[j][3], arr[i][0], arr[i][1], arr[j][0], arr[j][1]])
                    visit[i] = True
                    j += 1
                    break
                else:
                    j += 1
                    continue
            elif j == len(arr) - 1 and i + next + 1 < len(arr) - 1:
                j = i
                visit[i + next] = True
                next = next + 1
                continue
            else:
                # print(j, ":", arr[i+next][2],"-", arr[j][2])
                j += 1

        # print(next)
        next = 1
        i += 1
    #print(len(block_point))

    return block_point

def ImgtoPadding(img, psize) : # psize : padding size
    rows, cols = img.shape
    rows = rows + psize*2
    cols = cols + psize*2
    padding_crop_img = np.empty((rows, cols))

    for j in range(0, rows):
        for k in range(0, cols):
            if j >= psize and k >= psize and j < rows - psize and k < cols - psize:
                padding_crop_img[j][k] = img[j - psize][k - psize]
            else:
                padding_crop_img[j][k] = 255

    return padding_crop_img

def StartTableExtract(img_path):
    ####################################
    ############ main start ############
    ####################################
    #print(cv2.__version__)

    img_file = img_path
    img = cv2.imread(img_file, cv2.IMREAD_GRAYSCALE)
    rows, cols = img.shape  # 이미지의 가로, 세로 길이
    if rows > 2000 and cols > 2000: 
        img = cv2.resize(img, None, fx=0.4, fy=0.4, interpolation=cv2.INTER_AREA)
    rows, cols = img.shape  # 이미지의 가로, 세로 길이
    
    gray = cv2.adaptiveThreshold(~img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 15, -2)
    block_point = GetBlockList(gray, rows, cols)  # <-- 완성
    #print(block_point)

    #########여기서부턴 엑셀에 넣는 거 ####### 함수 아직 안만들었음
    ######################################
    wb = Workbook()
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    # double = Side(border_style="double", color="ff0000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # font = Font(b=True, color="FF0000")
    al = Alignment(horizontal="center", vertical="center")
    # wb.save("styled.xlsx")

    matching_col = ['A', 'B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF','AG','AH','AI','AJ','AK']
    img_list = []
    #
    # ws.merge_cells(start_row=2, start_column=4, end_row=8, end_column=6)
    numb=0
    for i in block_point:
        ws.merge_cells(start_row=i[1]+1, start_column=i[0]+1, end_row=i[3], end_column=i[2])
        # print(numb)
        # numb = numb + 1
        merge_string = matching_col[i[0]] + str(i[1] + 1) + ':' + matching_col[i[2] - 1] + str(i[3])
        style_range(ws, merge_string, border=border, alignment=al)

        cell = ws.cell(row=i[1] + 1, column=i[0] + 1)
        cell.border = border

        # # 자르는 이미지 범위 설정
        crop_img = img[int(i[5]) - 2 : int(i[7]) - 1, int(i[4]) + 3:int(i[6]) - 3]
        padding_crop_img = ImgtoPadding(crop_img, 10)
        crop_position = str(int(i[0])) + "/" + str(int(i[1])) + "/" + str(int(i[2])) + "/" + str(int(i[3]))
        StoreFormat = [padding_crop_img, crop_position]
        img_list.append(StoreFormat)

        # 이미지 저장
        #crop_name = "./OpenCV/res/image" + str(int(i[0])) + str(int(i[1])) + str(int(i[2])) + str(int(i[3])) + '.png'
        #
        #img_result = cv2.resize(padding_crop_img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        #cv2.imwrite(crop_name, img_result)
        # # print(crop_name)
        # cv2.imshow(crop_name, crop_img)

    cv2.waitKey(0)
    cv2.destroyAllWindows()
    wb.save("result.xlsx")
    return img_list

def StoretoExcel(inf) : # inf : [[학습결과, "x0/y0/x1/y2"]
    wb = load_workbook('./result.xlsx')
    ws = wb.active
    for i in inf:
        k = i[1].split("/")
        if i[0] is '':
            i[0] = 0
        ws.cell(row = int(k[1])+1, column= int(k[0])+1, value="{}".format(i[0]))
    wb.save("result.xlsx")

#StartTableExtract('../test-2.png') # for testing.
