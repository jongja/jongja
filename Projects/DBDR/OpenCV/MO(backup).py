import cv2
import numpy as np
import operator
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, named_styles, colors, Font
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

MAX_CORNERS = 100000

def GetBinaryImg(img) :
    return cv2.adaptiveThreshold(~img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 15, -2)


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws[cell_range.split(":")[0]]

    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment
        rows = ws[cell_range]

    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top

    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right

        if fill :
            for c in row:
                c.fill = fill

def openOperation(img, imgsize, select) : # 열림 연산 // select : 0(가로선), 1(세로선)
    # 이미지 크기에 따른 scale 값 지정
    if imgsize > 800:
        scale = 30
    elif imgsize > 400:
        scale = 20
    else:
        scale = 10

    size = imgsize / scale
    if select == 0 :
        window = np.ones((int(size), 1), np.uint8)
    elif select == 1 :
        window = np.ones((1, int(size)), np.uint8)

    # para1 : 이미지, para2 : 커널, para3 : erode 반복 횟수
    img = cv2.erode(img, window, iterations=1)
    # para1 : 이미지, para2 : 커널, para3 : dilate 반복 횟수
    img = cv2.dilate(img, window, iterations=1)
    return img

def CheckTopBottemLine(horizontal_img, x1, x2, y) :
    xpos_avg = int((x1 + x2) / 2)  # 사각형의 x1, x2의 평균

    check = 0 # check가 1이면 bottom line이 있는 것
    if horizontal_img[int(y)][xpos_avg] > 240 : # 240보다 크면 선이 있다고 판단한다.
        check = 1
    elif horizontal_img[int(y)+1][xpos_avg] > 240:
        check = 1
    elif horizontal_img[int(y)+2][xpos_avg] > 240:
        check = 1
    elif horizontal_img[int(y)-1][xpos_avg] > 240:
        check = 1
    elif horizontal_img[int(y)-2][xpos_avg] > 240:
        check = 1

    if check == 1 : # 선이 발견된 경우
        return True
    else :
        return False

def CheckLeftRightLine(vertical_img, y1, y2, x) :
    ypos_avg = int((y1 + y2) / 2)  # 사각형의 y1, y2의 평균

    check = 0 # check가 1이면 bottom line이 있는 것

    if vertical_img[ypos_avg][int(x)] > 240 : # 240보다 크면 선이 있다고 판단한다.
        check = 1
    elif vertical_img[ypos_avg][int(x) + 1] > 240:
        check = 1
    elif vertical_img[ypos_avg][int(x) + 2] > 240:
        check = 1
    elif vertical_img[ypos_avg][int(x) - 1] > 240:
        check = 1
    elif vertical_img[ypos_avg][int(x) - 2] > 240:
        check = 1

    if check == 1 : # 선이 발견된 경우
        return True
    else :
        return False

def PositionAlignment(img, lines) :
    for i in range(len(lines)):
        for j in range(len(lines)):
            if abs(lines[i][0][0] - lines[j][0][0]) < 10 and lines[i][0][0] != lines[j][0][0]:
                lines[j][0][0] = lines[i][0][0]

            if abs(lines[i][0][1] - lines[j][0][1]) < 10 and lines[i][0][1] != lines[j][0][1]:
                lines[j][0][1] = lines[i][0][1]

            if abs(lines[i][0][2] - lines[j][0][2]) < 10 and lines[i][0][2] != lines[j][0][2]:
                lines[j][0][2] = lines[i][0][2]

            if abs(lines[i][0][3] - lines[j][0][3]) < 10 and lines[i][0][3] != lines[j][0][3]:
                lines[j][0][3] = lines[i][0][3]

    for i in range(len(lines)) :
        cv2.line(img, (lines[i][0][0], lines[i][0][1]), (lines[i][0][2], lines[i][0][3]), (255, 0, 0), 3)

    return img

def SetIndexList(corner, offset, type) :
    list = []
    temp_list = []

    for i in corner:
        x, y = i.ravel()
        temp_list.append([x, y])

    temp_list.sort()

    if type == 0 : # type 0 -> horizontal list
        # column의 좌표만 추가 -> x 좌표만 추가한다
        for i in temp_list:
            x = i[0]

            if len(list) == 0:  # 비어있을 경우
                list.append([x - offset, x + offset, 100])  # x 좌표 추가(범위로 넣는다.)
            elif list[len(list) - 1][1] < x:  # x 좌표를 새로 넣어야 하는 경우
                list.append([x - offset, x + offset, 100])  # x 좌표 추가(범위로 넣는다.)

    elif type == 1 : # type 1 -> vertical list
        temp_list.sort(key=operator.itemgetter(1))
        # row 좌표만 추가 -> y 좌표만 추가한다
        for i in temp_list:
            y = i[1]

            if len(list) == 0:  # 비어있을 경우
                list.append([y - offset, y + offset, 100])  # y 좌표 추가(범위로 넣는다.)
            elif list[len(list) - 1][1] < y:  # y 좌표를 새로 넣어야 하는 경우
                list.append([y - offset, y + offset, 100])  # y 좌표 추가(범위로 넣는다.)

    list.sort()  # 오름 차순 정렬

    for i in range(len(list)):  # 좌표 매핑
        list[i][2] = i

    return list

def GetImprovedVertical(vertical, threshold = 80) : # threshold : r,θ 평면에서 몇개의 곡선이 한점에서 만났을 때 직선으로 판단할지에 대한 최소값
    vertical_lines = cv2.HoughLinesP(vertical, 1, np.pi / 180, threshold, 2, 11)
    vertical = PositionAlignment(vertical, vertical_lines)
    return vertical

def GetImprovedHorizontal(horizontal, threshold = 150) : # threshold : r,θ 평면에서 몇개의 곡선이 한점에서 만났을 때 직선으로 판단할지에 대한 최소값
    horizontal_lines = cv2.HoughLinesP(horizontal, 1, np.pi / 180, threshold, 2, 11)
    horizontal = PositionAlignment(horizontal, horizontal_lines)
    return horizontal

def GetMRPointList(horizontal, vertical, horizontal_list, vertical_list) :
    mask = ~horizontal + ~vertical

    corners = cv2.goodFeaturesToTrack(mask, MAX_CORNERS, 0.01, 10)
    existed_point = []  # [x좌표, y좌표, x 매핑좌표, y 매핑좌표]

    ##################### img의 모든 점을 실제 좌표와 맵핑 좌표로 저장해주기 #######################
    for i in corners:
        x, y = i.ravel()
        existed_point.append([x, y, 0, 0])
        cv2.circle(mask, (x, y), 5, 0, 1) # 그냥 보는 용도

    for i in range(len(existed_point)):
        for rangeCheck in horizontal_list:  # x 좌표 매핑 업데이트
            if existed_point[i][0] > rangeCheck[0] and existed_point[i][0] < rangeCheck[1]:  # 해당 x 범위 안에 있을 경우
                existed_point[i][2] = rangeCheck[2]
                break

        for rangeCheck in vertical_list:  # y 좌표 매핑 업데이트
            if existed_point[i][1] > rangeCheck[0] and existed_point[i][1] < rangeCheck[1]:  # 해당 y 범위 안에 있을 경우
                existed_point[i][3] = rangeCheck[2]

    # 오름차순 정렬
    existed_point.sort(key=operator.itemgetter(3))
    existed_point.sort(key=operator.itemgetter(2))
    #cv2.imshow('mask', mask)
    return existed_point

def GetBlockList(gray, rows, cols) :
    vertical = openOperation(gray, cols, 0)
    vertical = GetImprovedVertical(vertical, 80)
    # cv2.imshow("vertical", vertical)
    horizontal_list = GetList(vertical, 0)  # horizontal list -> img : vertical / type : 0 넣어주기
    xMaxSize = len(horizontal_list) - 1

    ##############################################################################
    horizontal = openOperation(gray, rows, 1)
    horizontal = GetImprovedHorizontal(horizontal, 150)
    # cv2.imshow("horizontal", horizontal)
    vertical_list = GetList(horizontal, 1)  # vertical list -> img : horizontal / type : 1 넣어주기
    yMaxSize = len(vertical_list) - 1

    ###################################################################
    ########### ExistedAllPoint로 작업해야함 !!!!!!!!!!!!!! #############
    ########### ExistedAllPoint로 block을 만들어야 함       #############
    ##################################################################
    ExistedAllPoint = GetMRPointList(horizontal, vertical, horizontal_list, vertical_list)

    block_point = []  # 블록 좌표가 저장되는 곳. [x1, y1, x2, y2] (매핑 좌표)
    temp = [0, 0, 0, 0, 0, 0, 0, 0]  # 임시로 저장하는 곳, 앞의 4개는 매핑 좌표, 뒤의 4개는 실제 좌표

    flag1 = 0  # 블록의 4개 좌표가 모두 구해질 경우 1 아니면 0 (블록이 x)
    flag2 = 0  # 블록의 4개 좌표가 모두 구해질 경우 1 아니면 0 (블록이 x)
    prev_y = 0

    # 시계방향
    for i in ExistedAllPoint:
        flag = 0
        bottom_top_flag = 0

        for compare in ExistedAllPoint:
            if bottom_top_flag == 1: break

            if i[2] == xMaxSize or i[3] == yMaxSize: break

            if prev_y > compare[3] and flag2 == 0:  # 이전 y 좌표보다 더 작은 y 좌표가 들어온 경우 x 좌표가 바뀌었다는 소리임 -> flag 초기화 해줘야함.
                flag1 = 0

            if i[3] == compare[3] and i[2] < compare[2] and flag1 == 0:  # 오른쪽 방향
                if CheckTopBottemLine(horizontal, i[0], compare[0], i[1]) == False:  # Top line check
                    break

                prev_y = compare[3]

                ## 매핑 좌표
                temp[0] = i[2]
                temp[1] = i[3]
                temp[2] = compare[2]
                temp[3] = 100

                ## 실제 좌표
                temp[4] = i[0]
                temp[5] = i[1]
                temp[6] = compare[0]
                temp[7] = 10000
                flag1 = 1

            if temp[2] == compare[2] and temp[1] < compare[3] and flag1 == 1:  # 밑 방향
                if CheckLeftRightLine(vertical, temp[5], compare[1], temp[6]) == False:  # Right line check
                    flag1 = 0
                    continue
                # 매핑 좌표
                temp[3] = compare[3]
                # 실제 좌표
                temp[7] = compare[1]
                flag2 = 1

            if flag1 == 1 and flag2 == 1:  # 세 개의 좌표를 구한 경우. 마지막 한 좌표 존재하는지 확인
                for flagCheck in ExistedAllPoint:
                    if flagCheck[2] == temp[0] and flagCheck[3] == temp[3]:  # 마지막 한 좌표 발견
                        if CheckTopBottemLine(horizontal, temp[4], temp[6], temp[7]) == False:  # Bottom line check
                            flag2 = 0  # 플래그 초기화
                            break

                        elif CheckLeftRightLine(vertical, temp[5], temp[7], temp[4]) == False:  # Left line check
                            flag1 = 0  # 플래그 초기화
                            flag2 = 0  # 플래그 초기화
                            bottom_top_flag = 1
                            break

                        flag = 1
                        flag1 = 0  # 플래그 초기화
                        flag2 = 0  # 플래그 초기화
                        block_temp = temp[:]
                        block_point.append(block_temp)  # 블록 리스트에 넣어준다.
                        break
                    elif flagCheck[2] > temp[0]:
                        flag2 = 0  # 플래그 초기화
                        break  # 뒤는 더이상 없으므로 for문 바로 빠져나옴

            if flag == 1:
                break  # 발견했으므로 for문을 빠져나가자 !
    return block_point

def GetList(img, type) :
    if type == 0 :
        corners_h = cv2.goodFeaturesToTrack(img, MAX_CORNERS, 0.01, 5)  # 코너 추출 함수
        return SetIndexList(corners_h, 10, type)
    else :
        corners_v = cv2.goodFeaturesToTrack(img, MAX_CORNERS, 0.01, 5)  # 코너 추출 함수
        return SetIndexList(corners_v, 10, type)

def ImgtoPadding(img, psize) : # psize : padding size
    rows, cols = img.shape
    rows = rows + psize*2
    cols = cols + psize*2
    padding_crop_img = np.empty((rows, cols))

    for j in range(0, rows):
        for k in range(0, cols):
            if j >= psize and k >= psize and j < rows - psize and k < cols - psize:
                padding_crop_img[j][k] = img[j - psize][k - psize]
            else:
                padding_crop_img[j][k] = 255

    return padding_crop_img

def StartTableExtract(img_path) :
    ####################################
    ############ main start ############
    ####################################
    img_file = img_path
    img = cv2.imread(img_file, cv2.IMREAD_GRAYSCALE)
    #img = cv2.resize(img, None, fx=0.4, fy=0.4, interpolation=cv2.INTER_AREA)
    rows, cols = img.shape # 이미지의 가로, 세로 길이
    gray = GetBinaryImg(img)
    block_point = GetBlockList(gray, rows, cols) # <-- 완성
    #########여기서부턴 엑셀에 넣는 거 ####### 함수 아직 안만들었음
    ######################################
    wb = Workbook()
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    #double = Side(border_style="double", color="ff0000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #font = Font(b=True, color="FF0000")
    al = Alignment(horizontal="center", vertical="center")
    #wb.save("styled.xlsx")

    matching_col = ['A', 'B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF','AG','AH','AI','AJ','AK']
    img_list = []

    for i in block_point :
        ws.merge_cells(start_row=i[1]+1, start_column=i[0]+1, end_row=i[3], end_column=i[2])

        merge_string = matching_col[i[0]] + str(i[1]+1) + ':' + matching_col[i[2]-1] + str(i[3])
        style_range(ws, merge_string, border=border, alignment=al)

        cell = ws.cell(row = i[1]+1, column = i[0]+1)
        cell.border = border

        # 자르는 이미지 범위 설정
        crop_img = img[int(i[5]) - 2:int(i[7]), int(i[4]) + 3:int(i[6]) - 3]
        padding_crop_img = ImgtoPadding(crop_img, 5)

        #crop_img = GetBinaryImg(crop_img)
        crop_position = str(int(i[0])) + "/" + str(int(i[1]))+ "/"  + str(int(i[2])) + "/"  + str(int(i[3]))
        StoreFormat = [padding_crop_img, crop_position]
        img_list.append(StoreFormat)

        # 이미지 저장
        crop_name = "./OpenCV/res/image" + str(int(i[0])) + str(int(i[1])) + str(int(i[2])) + str(int(i[3])) + '.png'
        #
        # # print(crop_name)
        img_result = cv2.resize(padding_crop_img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        cv2.imwrite(crop_name, img_result)
        #cv2.imshow(crop_name, crop_img)

    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    wb.save("result.xlsx")
    return img_list

def StoretoExcel(inf) : # inf : [[학습결과, "x0/y0/x1/y2"]
    wb = load_workbook('result.xlsx')
    ws = wb.active
    for i in inf:
        k = i[1].split("/")
        if i[0] is '':
            i[0] = 0
        ws.cell(row = int(k[1])+1, column= int(k[0])+1, value="{}".format(i[0]))
    wb.save("result.xlsx")
